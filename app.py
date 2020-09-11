"""
Hydra project application

Using Beeware, this app produces a simple Gui that allows 
users to enter in a series of search terms for the NHS Hydra pfam
portal and returns a set of transcripts for each term, then saves them  
into an excel spreadsheet. 
"""
import toga
from toga.style import Pack
from toga.style.pack import COLUMN, ROW
import time
from selenium import webdriver
#import geckodriver_autoinstaller #is this being used? Has it been installed and is no longer needed? Problem, geckodriver has to be installed before running app. 
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import json, requests, sys, pprint
import openpyxl
from openpyxl.styles import Font
import urllib.request, urllib
import re

#geckodriver_autoinstaller.install()  - 

class Hydra(toga.App):
    
    def startup(self):
        self.terms = []
        main_box = toga.Box(style=Pack(direction=COLUMN))

        term_label = toga.Label(
            'Enter term(s): ',
            style=Pack(padding=(0, 5))
        )
        self.term_input = toga.TextInput(style=Pack(flex=1))

        term_box = toga.Box(style=Pack(direction=ROW, padding=5))
        term_box.add(term_label)
        term_box.add(self.term_input)

        term_button = toga.Button(
            'Add term',
            on_press= self.term_list,
            style=Pack(padding=5)
        )
        
        fetch_button = toga.Button(
            'Fetch transcripts',
            on_press= self.hydra_fetch,
            style=Pack(padding=5)
        )
        
        main_box.add(term_box)
        main_box.add(term_button)
        main_box.add(fetch_button)

        self.main_window = toga.MainWindow(title=self.formal_name)
        self.main_window.content = main_box
        self.main_window.show()


    def term_list(self, widget):
        if self.term_input.value: 
            self.terms.append(self.term_input.value)
        else: 
            print('Please enter a term')

        self.term_input.value = ''
        return self.terms
    
    
    def hydra_fetch(self,widget):
        ''' Accepts user input and returns a list of terms '''
        ''' Input (list); Return list'''
    # **use the term "Ig_4" to bug test this program; it returns a short list of results

        self.main_window.info_dialog(
            'Click to fetch transcripts, otherwise enter more terms',
            "Terms, {}".format(self.terms)
        )

        self.main_window.info_dialog(
                'Click to Fetch ....',
                ''
        )
        # Open Firefox in headless(invisible), mode
        options = webdriver.FirefoxOptions()
        options.add_argument('--headless')
        self.browser = webdriver.Firefox(options=options)


        # Enter search terms on the Hydra homepage and get the geneids. 
        pfam_webportal = 'https://research.nhgri.nih.gov/hydra/pfam/'
        self.browser.get(pfam_webportal)
        self.search_geneid(self.terms,self.browser)  

        # Make urls from geneids using gb_link, then download JSON data and strip out transcripts
        
        self.transcripts =[]
        count = 0 
        for self.geneid in self.geneids:
            self.gb_link(self.geneid, self.browser) 
            self.transcripts_data(self.jbrowse_url, self.browser)
            self.transcripts.append(self.transcript_set)
            if count % 3 ==0: 
                print('waiting ...')
                time.sleep(60)
            count +=1

        self.browser.close()

        #Put all terms and their matching geneids and transcripts into an excel file.
        self.manage_excel(self.terms, self.geneids, self.transcripts)

        self.main_window.info_dialog(
            'Done',
            'Transcripts placed in Excel file'
        )


    def search_geneid(self,terms, browser): 
        '''Input a search term and return a geneid '''
        '''Input(list), Return(list)'''

        # Find the keyword field and enter the term 
        self.keyword_box = '//div[3]/form/table[4]/tbody/tr[2]/td[5]/input[1]'

        self.keywordElem = WebDriverWait(self.browser, 60).until(lambda x:x.find_element_by_xpath(self.keyword_box))
        self.keywordElem.click()
        print('Found keywordElem with that xpath!')

        for self.term in self.terms: 

            self.keywordElem.send_keys(self.term)

            self.button = '//div[3]/form/table[4]/tbody/tr[2]/td[5]/input[2]'
            self.buttonElm = WebDriverWait(self.browser, 60).until(lambda x:x.find_element_by_xpath(self.button))
            self.buttonElm.click()
            print ('Found buttonElm with that xpath')

            #Return geneid links produced by search
            self.geneids =[]
            result_link = '//div[3]/table[3]/tbody/tr/td[1]/form/table[1]/tbody/tr[*]/td[2]/a'

            for link in self.browser.find_elements_by_xpath(result_link):
                self.geneid = link.text
                # Split the geneid to provide id to make gene sequence page url.
                self.geneids.append((self.geneid.split("="))[0]) 
            
    
    def transcripts_data(self,url,browser): 
        '''Download the JSON data as a file from the JSON browser '''
        '''Input: url(string) # Output: transcripts(list)''' 

        self.transcript_set =[0]
        response = requests.get(url)
        # raise_for_status lets you know if the file downloaded sucessfully
        response.raise_for_status() 
        # Load JSON data into a Python variable.
        #transcripts_data = json.loads(response.text)
        #transcripts_text_file = transcripts_data['intervals']
        
        #regex expression
        # 't\d+aep' stands for find the letter 't' and a number from 0-9 '\d' plus the letters 'aep'
        pattern = re.compile('t\d+aep')
        self.transcript_set = set(pattern.findall(response.text))
        

    def gb_link(self,geneid, browser): 
        '''Takes a gene id and clicks through _View Gene in Genome Browser_ webpage to jbrowser.'''
        '''Input: genid(string), Output: url (string)'''
        
        self.url = 'https://research.nhgri.nih.gov/hydra/genewiki/gene_page.cgi?gene=' + self.geneid
        self.browser.get(self.url)

        #Click the 'View Gene in Genome Browser link' - it seems to be the only way to get the correct json file
        link_to_genome = WebDriverWait(self.browser, 60).until(lambda x:x.find_element_by_link_text('View Gene in Genome Browser'))
        link_to_genome.click()
        print ('Found View_gene_elem with that xpath, clicked')
        
        # After clicking the button, store the window handle variable of the view_in_broswer page
        #Then wait for the jbrowse page to load and store that window handle variable 
        view_in_browser_page = self.browser.window_handles[0] #should this be stored before clicking? 
        print('View in browser window handle stored')
        jbrowser_page =browser
        checkbox = '//div[2]/div[4]/div[5]/div[2]/div/div/label[2]/input'
        #time.sleep(5) # putting this sleep here allowed me to get 6 transcripts before it choked
        jbrowser_page = ''
        self.browser.switch_to.window(jbrowser_page)
        print ("switched to jbrowser page")

        # This is crucial!!! If the browser does not have time to load the second page, 
        # the handle will not be reset to the correct page. Sleep time may need to be lengthened 

        checkbox = '//div[2]/div[4]/div[5]/div[2]/div/div/label[2]/input'
        JulianoCheckbox = WebDriverWait(self.browser, 60).until(lambda x:x.find_element_by_xpath(checkbox))
        JulianoCheckbox.click()
        print('clicked Juliano checkbox')
    
        self.browser.refresh()
        self.short_geneid = geneid.split('.')
        self.jbrowse_url = 'https://research.nhgri.nih.gov/hydra/jbrowse/data/tracks/aepLRv2_splign/'\
                      + self.short_geneid[0] + '/trackData.json'
        print ("jbrowse url", self.jbrowse_url)
        self.browser.close()
        print ('Jbrowser close')
        #self.browser.switch_to.window(view_in_browser_page)
        #print('switched back to view in browser for next gene url')


    def manage_excel(self,terms, geneid, transcripts): 
        '''Excel file with a title based on a website and the date'''

        title = 'Pfam_Dom'   
        subtitle = 'Pfam domains in predicted Hydra proteins'
        web_address = 'https://research.nhgri.nih.gov/hydra/pfam/'
        datestamp = time.strftime('%Y/%m/%d_%H'+':'+'%M')

        title = title.replace(' ','')
        wb = openpyxl.Workbook() # Create a blank workbook
        wb.sheetnames # It starts with one sheet
        sheet = wb.active
        sheet.title
        sheet.title = title 

        #Create fonts
        title_style = Font(size=24, bold=True) 
        term_style = Font(size =16, bold = True)
        term_sub_style = Font(size =16)
        geneid_style = Font(size=14, bold = True)
        transcript_style = Font(size=12, bold = True)

        #Format spreadsheet
        sheet['A1'] = title
        sheet['A1'].font = title_style
        sheet['A2'] = subtitle
        sheet['A3'] = web_address
        sheet['A5'] = datestamp 
        #blank_cell= " "
        
        row_num = 6
        col_num = 1

        row_num += 2
        sheet.cell(row=row_num, column=col_num).font = term_style
        sheet.cell(row=row_num, column=col_num).value = "TERM"


        if len(terms) <= 1:
            term = str(terms)

            row_num += 1
            sheet.cell(row=row_num, column=col_num).font = term_sub_style
            sheet.cell(row=row_num, column=col_num).value = term

            row_num += 1
            sheet.cell(row=row_num, column=col_num).font = geneid_style
            sheet.cell(row=row_num, column=col_num).value = "Geneids"
            col_num +=1
            sheet.cell(row=row_num, column=col_num).font = transcript_style
            sheet.cell(row=row_num, column=col_num).value = "transcripts"
            col_num -=1 

            self.geneid_list = self.geneids
            self.transcript_sets = self.transcripts

            if len(self.geneid_list) <=1: 
                row_num += 1
                col_num = 1
                sheet.cell(row=row_num, column=col_num).value = self.geneid_list
                for self.transcript_set in self.transcript_sets: 
                    for self.transcript in self.transcript_set: 
                        col_num +=1
                        sheet.cell(row=row_num, column=col_num).value = self.transcript

            else: 
                for self.geneid, self.transcript_set in zip(self.geneid_list, self.transcript_sets):
                    row_num += 1
                    col_num = 1
                    sheet.cell(row=row_num, column=col_num).value = self.geneid
    
                    for self.transcript in self.transcript_set: 
                        col_num +=1
                        sheet.cell(row=row_num, column=col_num).value = self.transcript

        else: 

            for self.term, self.geneid_list, self.transcript_set in zip(self.terms, self.geneid_lists, self.transcript_sets): 
                # Search terms should start at cell A7, row

                col_num = 1
                row_num += 1
                sheet.cell(row=row_num, column=col_num).font = term_sub_style
                sheet.cell(row=row_num, column=col_num).value = self.term

                row_num += 1
                sheet.cell(row=row_num, column=col_num).font = geneid_style
                sheet.cell(row=row_num, column=col_num).value = "Geneids"
                col_num +=1
                sheet.cell(row=row_num, column=col_num).font = transcript_style
                sheet.cell(row=row_num, column=col_num).value = "transcripts"
                col_num -=1 

                for self.geneid, self.transcript_set in zip(self.geneid_list, self.transcript_sets):
                    row_num += 1
                    col_num = 1
                    sheet.cell(row=row_num, column=col_num).value = self.geneid
                
                    for self.transcript in self.transcript_set: 
                        col_num +=1
                        sheet.cell(row=row_num, column=col_num).value = self.transcript

        wb.save( 'User\Documents\hydra-app'+title+'_'+ datestamp+'.xlsx' ) 


def main():
    return Hydra()
